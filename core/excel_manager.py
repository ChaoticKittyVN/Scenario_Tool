# core/excel_manager.py
"""
Excel读取器模块
提供统一的Excel文件读取和DataFrame处理功能

职责：
1. ExcelFileManager - 管理Excel文件的读取、缓存和工作表访问
2. DataFrameProcessor - 处理DataFrame数据的提取、过滤和验证
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional, Callable, Union
from functools import wraps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.logger import get_logger
from core.constants import SheetName, ColumnName, Marker
logger = get_logger(__name__)

# ==================== 异常定义 ====================
class ExcelManagerError(Exception):
    """Excel读取基础异常"""
    def __init__(self, message: str, original_error: Exception = None):
        super().__init__(message)
        self.original_error = original_error
        self.message = message


class ExcelFileNotFoundError(ExcelManagerError):
    """文件不存在异常"""
    pass


class ExcelFormatError(ExcelManagerError):
    """文件格式错误"""
    pass


class ExcelDataError(ExcelManagerError):
    """数据内容错误"""
    pass


class ExcelWriteError(ExcelManagerError):
    """写入错误"""
    pass

# ==================== 错误处理装饰器 ====================
def handle_excel_operation(func):
    """Excel操作统一错误处理装饰器"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except FileNotFoundError as e:
            logger.error(f"文件不存在: {e}")
            raise ExcelFileNotFoundError(f"Excel文件不存在: {e}", e)
        except pd.errors.EmptyDataError as e:
            logger.error(f"Excel文件为空: {e}")
            raise ExcelFormatError(f"Excel文件为空或格式错误: {e}", e)
        except pd.errors.ParserError as e:
            logger.error(f"Excel解析失败: {e}")
            raise ExcelFormatError(f"Excel文件解析失败: {e}", e)
        except PermissionError as e:
            logger.error(f"文件访问权限不足: {e}")
            raise ExcelFileNotFoundError(f"无法访问Excel文件: {e}", e)
        except ExcelManagerError:
            # 如果是我们自定义的异常，直接抛出
            raise
        except Exception as e:
            logger.error(f"处理Excel时发生未知错误: {e}", exc_info=True)
            raise ExcelManagerError(f"处理Excel失败: {e}", e)
    return wrapper

# ==================== Excel文件管理器 ====================
class ExcelFileManager:
    """
    Excel文件管理器
    负责Excel文件的读取、缓存和工作表管理
    """
    
    def __init__(self, cache_enabled: bool = True):
        """
        初始化Excel文件管理器
        
        Args:
            cache_enabled: 是否启用文件缓存
        """
        self._file_cache: Dict[Path, Dict[str, pd.DataFrame]] = {}
        self.cache_enabled = cache_enabled

    @handle_excel_operation
    def load_excel(self, file_path: Path) -> Dict[str, pd.DataFrame]:
        """
        加载Excel文件的所有工作表
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict[str, pd.DataFrame]: 工作表名到DataFrame的映射
            
        Raises:
            ExcelFileNotFoundError: 文件不存在
            ExcelFormatError: 文件格式错误
            ExcelManagerError: 其他读取错误
        """
        if not file_path.exists():
            logger.error(f"Excel文件不存在: {file_path}")
            raise ExcelFileNotFoundError(f"Excel文件不存在: {file_path}")
        
        # 检查缓存
        if file_path in self._file_cache and self.cache_enabled:
            logger.debug(f"从缓存加载Excel文件: {file_path}")
            return self._file_cache[file_path]
        
        logger.info(f"加载Excel文件: {file_path}")
        try:
            # 读取所有工作表，所有列作为字符串类型处理
            data = pd.read_excel(file_path, sheet_name=None, dtype=str)
            
            # 清理数据：将NaN转换为空字符串
            for sheet_name, df in data.items():
                data[sheet_name] = df.fillna("")
            
            if self.cache_enabled:
                self._file_cache[file_path] = data
            
            logger.debug(f"文件加载成功: {file_path}, 工作表数: {len(data)}")
            return data
            
        except pd.errors.EmptyDataError as e:
            logger.error(f"Excel文件为空: {file_path}")
            raise ExcelFormatError(f"Excel文件为空: {file_path}", e)
        except Exception as e:
            logger.error(f"读取Excel文件失败: {file_path}", exc_info=True)
            raise ExcelManagerError(f"读取Excel文件失败: {file_path}", e)
    
    def get_sheet(self, file_path: Path, sheet_name: str) -> pd.DataFrame:
        """
        获取指定工作表的DataFrame
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            pd.DataFrame: 指定工作表的DataFrame，如果不存在则返回空DataFrame
            
        Raises:
            ExcelManagerError: 文件读取错误
        """
        try:
            data = self.load_excel(file_path)
            df = data.get(sheet_name, pd.DataFrame())
            
            if df.empty:
                logger.warning(f"工作表不存在或为空: {file_path} -> {sheet_name}")
            else:
                logger.debug(f"获取工作表: {file_path} -> {sheet_name}, 形状: {df.shape}")

            return df
        except ExcelManagerError:
            # 重新抛出已有的ExcelManagerError
            raise
        except Exception as e:
            logger.error(f"获取工作表失败: {file_path} -> {sheet_name}", exc_info=True)
            raise ExcelManagerError(f"获取工作表失败: {file_path} -> {sheet_name}", e)

    def get_sheet_names(self, file_path: Path) -> List[str]:
        """
        获取Excel文件的所有工作表名称
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            List[str]: 工作表名称列表
            
        Raises:
            ExcelManagerError: 文件读取错误
        """
        try:
            data = self.load_excel(file_path)
            return list(data.keys())
        except ExcelManagerError:
            raise
        except Exception as e:
            logger.error(f"获取工作表名称失败: {file_path}", exc_info=True)
            raise ExcelManagerError(f"获取工作表名称失败: {file_path}", e)

    @handle_excel_operation
    def save_excel(self, 
                   file_path: Path, 
                   data: Dict[str, pd.DataFrame],
                   engine: str = 'openpyxl',
                   **kwargs) -> bool:
        """
        保存数据到Excel文件
        
        Args:
            file_path: 保存路径
            data: {工作表名: DataFrame} 的字典
            engine: 写入引擎 ('openpyxl' 或 'xlsxwriter')
            **kwargs: 传递给 pd.ExcelWriter 的额外参数
            
        Returns:
            bool: 是否保存成功
            
        Raises:
            ExcelWriteError: 写入失败
        """
        try:
            # 确保目录存在
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            logger.info(f"保存Excel文件: {file_path}")
            
            # 使用ExcelWriter保存多个工作表
            with pd.ExcelWriter(file_path, engine=engine, **kwargs) as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.debug(f"写入工作表: {sheet_name}, 形状: {df.shape}")
            
            logger.info(f"文件保存成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {file_path}", exc_info=True)
            raise ExcelWriteError(f"保存Excel文件失败: {file_path}", e)
    
    @handle_excel_operation
    def save_single_sheet(self,
                         file_path: Path,
                         df: pd.DataFrame,
                         sheet_name: str = 'Sheet1',
                         **kwargs) -> bool:
        """
        保存单个工作表到Excel文件
        
        Args:
            file_path: 保存路径
            df: 要保存的DataFrame
            sheet_name: 工作表名称
            **kwargs: 传递给 DataFrame.to_excel 的参数
            
        Returns:
            bool: 是否保存成功
        """
        try:
            # 确保目录存在
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            logger.info(f"保存Excel文件（单工作表）: {file_path}")
            df.to_excel(file_path, sheet_name=sheet_name, index=False, **kwargs)
            
            logger.info(f"文件保存成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {file_path}", exc_info=True)
            raise ExcelWriteError(f"保存Excel文件失败: {file_path}", e)

    def reload_file(self, file_path: Path):
        """
        重新加载文件（清除缓存并重新读取）

        Args:
            file_path: Excel文件路径
        """
        if file_path in self._file_cache:
            del self._file_cache[file_path]
            logger.info(f"清除文件缓存: {file_path}")

        # 重新加载文件
        self.load_excel(file_path)

    def clear_cache(self):
        """清除所有文件缓存"""
        self._file_cache.clear()
        logger.info("清除所有Excel文件缓存")


# ==================== DataFrame处理器 ====================
class DataFrameProcessor:
    """
    数据框处理器
    负责DataFrame数据的提取、过滤和验证
    """
    
    def __init__(self, config=None):
        """
        初始化数据框处理器
        
        Args:
            config: 应用配置对象
        """
        self.config = config
    
    @handle_excel_operation
    def extract_valid_rows(self, df: pd.DataFrame, sheet_name: str = "") -> pd.DataFrame:
        """
        提取有效行数据（基于END标记和忽略规则）
        
        Args:
            df: 输入的DataFrame
            sheet_name: 工作表名称（用于日志）
            
        Returns:
            pd.DataFrame: 有效行的DataFrame
            
        Raises:
            ExcelDataError: 数据格式错误
        """
        try:
            if df.empty:
                logger.warning(f"数据框为空，工作表: {sheet_name}")
                return df
            
            # 查找END标记位置
            end_index = self.find_marker_position(df, ColumnName.NOTE.value, Marker.END.value)
            if end_index == -1:
                logger.warning(f"未找到END标记，工作表: {sheet_name}")
                return pd.DataFrame()
            
            # 提取END标记之前的行
            valid_df = df.iloc[:end_index].copy()
            
            # 应用忽略规则（如果配置存在）
            if self.config and hasattr(self.config, 'processing'):
                valid_df = self._apply_ignore_rules(valid_df)
            
            logger.info(f"提取有效行: {sheet_name}, 原始行数: {len(df)}, 有效行数: {len(valid_df)}")
            return valid_df
            
        except Exception as e:
            logger.error(f"提取有效行失败: {sheet_name}", exc_info=True)
            raise ExcelDataError(f"提取有效行失败: {sheet_name}", e)
    
    def _apply_ignore_rules(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        应用忽略规则过滤数据
        
        Args:
            df: 输入的DataFrame
            
        Returns:
            pd.DataFrame: 过滤后的DataFrame
        """
        if not self.config.processing.ignore_mode:
            return df

        ignore_column = ColumnName.IGNORE.value
        if ignore_column not in df.columns:
            return df

        # 过滤掉标记为忽略的行
        mask = ~df[ignore_column].isin(self.config.processing.ignore_words)
        filtered_df = df[mask].reset_index(drop=True)

        ignored_count = len(df) - len(filtered_df)
        if ignored_count > 0:
            logger.info(f"忽略 {ignored_count} 行数据")

        return filtered_df
    
    def get_column_data(self, df: pd.DataFrame, column_name: str, default_value: Any = "") -> pd.Series:
        """
        安全获取列数据（处理列不存在的情况）
        
        Args:
            df: 输入的DataFrame
            column_name: 列名
            default_value: 列不存在时的默认值
            
        Returns:
            pd.Series: 列数据
        """
        if column_name in df.columns:
            return df[column_name]
        else:
            logger.warning(f"列不存在: {column_name}，返回默认值")
            return pd.Series([default_value] * len(df))
    
    def find_marker_position(self, df: pd.DataFrame, marker_column: str, marker_value: str) -> int:
        """
        查找标记位置（如END标记）
        
        Args:
            df: 输入的DataFrame
            marker_column: 标记列名
            marker_value: 标记值
            
        Returns:
            int: 标记位置索引，未找到返回-1
        """
        try:
            if marker_column not in df.columns:
                logger.debug(f"标记列不存在: {marker_column}")
                return -1
            
            # 查找包含标记值的行
            matches = df[df[marker_column] == marker_value]
            if not matches.empty:
                position = matches.index[0]
                logger.debug(f"找到标记 '{marker_value}' 在位置 {position}")
                return position
                
            logger.debug(f"未找到标记 '{marker_value}'")
            return -1
            
        except Exception as e:
            logger.error(f"查找标记位置失败: {marker_column}={marker_value}", exc_info=True)
            # 这里不抛出异常，返回-1表示未找到
            return -1
    
    def validate_dataframe(self, df: pd.DataFrame, required_columns: List[str]) -> bool:
        """
        验证DataFrame是否包含必需的列
        
        Args:
            df: 输入的DataFrame
            required_columns: 必需的列名列表
            
        Returns:
            bool: 是否验证通过
        """
        if df.empty:
            logger.warning("数据框为空")
            return False

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.warning(f"缺少必需列: {missing_columns}")
            return False

        return True
    
    def has_valid_data(self, value: Any) -> bool:
        """检查值是否有效（非空且非NaN）"""
        if pd.isna(value):
            return False
        if value == "":
            return False
        if value is None:
            return False
        return True
    
    def extract_parameters(self, row_data: pd.Series, needed_params: List[str]) -> Dict[str, Any]:
        """
        从行数据中提取指定参数
        
        Args:
            row_data: pandas Series，一行的所有数据
            needed_params: 需要的参数名列表
            
        Returns:
            Dict[str, Any]: 提取的参数字典
        """
        row_dict = row_data.to_dict()
        params = {}
        
        for param_name in needed_params:
            if param_name in row_dict:
                value = row_dict[param_name]
                if self.has_valid_data(value):
                    params[param_name] = value
        
        return params
    
    def extract_generator_params(self, row_data: pd.Series, generator_param_map: Dict) -> Dict:
        """
        为所有生成器提取参数
        
        Args:
            row_data: pandas Series，一行的所有数据
            generator_param_map: 生成器到参数的映射
            
        Returns:
            Dict: 生成器到参数的映射
        """
        generator_params = {}

        for generator, needed_params in generator_param_map.items():
            params = self.extract_parameters(row_data, needed_params)
            if params:
                generator_params[generator] = params
        
        return generator_params
    
    def extract_mapping_columns(self, df: pd.DataFrame, key_column: str, value_column: str) -> Dict[str, str]:
        """
        从DataFrame中提取两列构建映射字典
        
        Args:
            df: 输入的DataFrame
            key_column: 键列名
            value_column: 值列名
            
        Returns:
            Dict[str, str]: 映射字典
        """
        mapping = {}
        
        if key_column not in df.columns or value_column not in df.columns:
            logger.warning(f"缺少必需的列: {key_column} 或 {value_column}")
            return mapping
        
        for _, row in df.iterrows():
            key = row[key_column]
            value = row[value_column]
            
            if self.has_valid_data(key) and self.has_valid_data(value):
                mapping[str(key)] = str(value)

        return mapping
    
    def extract_param_names(self, df: pd.DataFrame, param_column: str = "ExcelParam") -> List[str]:
        """
        从DataFrame中提取参数名列表（用于填写参数表）
        
        Args:
            df: 输入的DataFrame
            param_column: 参数列名
            
        Returns:
            List[str]: 去重后的参数名列表
        """
        if param_column not in df.columns:
            logger.warning(f"参数列不存在: {param_column}")
            return []
        
        param_names = set()
        for _, row in df.iterrows():
            param = row[param_column]
            if self.has_valid_data(param):
                param_names.add(str(param))

        return sorted(list(param_names))
    
    def extract_columns_for_statistics(self, df: pd.DataFrame, columns: List[str], 
                                    keep_all_rows: bool = False) -> Dict[str, pd.Series]:
        """
        提取用于统计的列
        
        Args:
            df: 输入的DataFrame
            columns: 要提取的列名列表
            keep_all_rows: 是否保持所有行（用None填充无效值）
            
        Returns:
            Dict[str, pd.Series]: 列名到Series的映射
        """
        result = {}

        for column in columns:
            if column in df.columns:
                if keep_all_rows:
                    # 保持所有行，无效值填充为None
                    series = df[column].apply(
                        lambda x: str(x).strip() if self.has_valid_data(x) else None
                    )
                else:
                    # 只保留有效数据的行
                    series = df[column][df[column].apply(self.has_valid_data)]
                result[column] = series
            else:
                logger.warning(f"统计列不存在: {column}")
                result[column] = pd.Series(dtype=object)
        
        return result


# ==================== Excel写入器（高级功能） ====================
class ExcelWriter:
    """
    Excel写入器（高级功能）
    提供更复杂的Excel写入功能，如样式、公式、命名区域等
    基于openpyxl实现
    """
    
    def __init__(self):
        """初始化Excel写入器"""
        pass
    
    @handle_excel_operation
    def update_parameter_sheet(self,
                              file_path: Path,
                              sheet_name: str,
                              parameter_data: Dict[str, List[str]]) -> bool:
        """
        更新参数表工作表（保留样式和公式）
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（如"参数表"）
            parameter_data: 参数数据 {参数类型: [参数值列表]}
            
        Returns:
            bool: 是否更新成功
        """
        try:
            logger.info(f"更新参数表: {file_path} -> {sheet_name}")

            # 加载工作簿
            wb = load_workbook(file_path)
            
            # 获取或创建工作表
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # 创建居中对齐样式
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入参数数据
            for col_idx, (param_type, param_values) in enumerate(parameter_data.items(), 1):
                # 写入表头
                ws.cell(row=1, column=col_idx, value=param_type)
                ws.cell(row=1, column=col_idx).alignment = center_alignment
                
                # 写入参数值
                for row_idx, value in enumerate(param_values, 2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
            
            # 保存工作簿
            wb.save(file_path)
            logger.info(f"参数表更新成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"更新参数表失败: {file_path}", exc_info=True)
            raise ExcelWriteError(f"更新参数表失败: {file_path}", e)
    
    @handle_excel_operation
    def create_validation_template(self,
                                  output_path: Path,
                                  sheet_name: str,
                                  columns: List[str]) -> bool:
        """
        创建数据验证模板
        
        Args:
            output_path: 输出文件路径
            sheet_name: 工作表名称
            columns: 列名列表
            
        Returns:
            bool: 是否创建成功
        """
        try:
            logger.info(f"创建验证模板: {output_path}")
            
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 写入表头
            for col_idx, column_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=column_name)
            
            # 保存工作簿
            wb.save(output_path)
            logger.info(f"验证模板创建成功: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"创建验证模板失败: {output_path}", exc_info=True)
            raise ExcelWriteError(f"创建验证模板失败: {output_path}", e)


# ==================== 快捷函数 ====================
def create_excel_manager(cache_enabled: bool = True) -> ExcelFileManager:
    """创建Excel文件管理器实例"""
    return ExcelFileManager(cache_enabled=cache_enabled)


def create_dataframe_processor(config=None) -> DataFrameProcessor:
    """创建DataFrame处理器实例"""
    return DataFrameProcessor(config)


def create_excel_writer() -> ExcelWriter:
    """创建Excel写入器实例"""
    return ExcelWriter()


# ==================== 导出 ====================
__all__ = [
    'ExcelManagerError',
    'ExcelFileNotFoundError',
    'ExcelFormatError',
    'ExcelDataError',
    'ExcelWriteError',
    'ExcelFileManager',
    'DataFrameProcessor',
    'ExcelWriter',
    'create_excel_manager',
    'create_dataframe_processor',
    'create_excel_writer',
]


ExcelManager = ExcelFileManager
DataProcessor = DataFrameProcessor