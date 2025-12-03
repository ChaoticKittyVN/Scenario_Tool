import pandas as pd
from pathlib import Path
from typing import Dict, List

from core.logger import get_logger
from core.constants import SheetName, ColumnName, Marker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .excel_decorators import handle_excel_operation
from .excel_exceptions import ExcelWriteError

logger = get_logger(__name__)


# ==================== Excel写入器（高级功能） ====================
class ExcelWriter:
    """
    Excel写入器（高级功能）
    提供更复杂的Excel写入功能，如样式、公式、命名区域等
    基于openpyxl实现
    """
    
    def __init__(self):
        """初始化Excel写入器"""
        pass
    
    @handle_excel_operation
    def create_validation_template(self,
                                output_path: Path,
                                sheet_name: str,
                                columns: List[str]) -> bool:
        """
        创建数据验证模板
        """
        logger.info(f"创建验证模板: {output_path}")
        
        # 输入验证
        if not output_path:
            raise ExcelWriteError("输出路径不能为空")
        
        if not sheet_name:
            sheet_name = "Sheet1"  # 提供默认值
            logger.warning(f"工作表名称为空，使用默认值: {sheet_name}")
        
        if not columns:
            columns = []  # 空列表，只创建空的工作表
            logger.warning("列名列表为空，将创建空工作表")
        
        # 确保输出目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 创建工作簿
        wb = Workbook()
        
        # 安全地获取或创建工作表
        try:
            # 尝试获取活动工作表
            ws = getattr(wb, 'active', None)
            
            if ws is None:
                # 如果没有活动工作表，创建一个
                logger.debug("工作簿没有活动工作表，创建新工作表")
                ws = wb.create_sheet(title=sheet_name)
            else:
                # 设置工作表名称
                try:
                    ws.title = sheet_name
                except Exception as name_error:
                    logger.warning(f"无法设置工作表名称 '{sheet_name}': {name_error}")
                    # 继续使用默认名称
        except Exception as e:
            logger.error(f"处理工作表时出错: {e}")
            raise ExcelWriteError(f"无法处理Excel工作表: {e}")
        
        # 写入表头（如果提供了列名）
        if columns:
            try:
                for col_idx, column_name in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = column_name if column_name else f"Column{col_idx}"
                
                logger.debug(f"成功写入 {len(columns)} 列表头")
            except Exception as header_error:
                logger.error(f"写入表头失败: {header_error}")
                # 不因表头写入失败而终止整个操作
                # 继续保存文件，只是没有表头
        
        # 保存文件
        try:
            wb.save(output_path)
            logger.info(f"验证模板创建成功: {output_path}")
            return True
        except Exception as save_error:
            logger.error(f"保存文件失败: {save_error}")
            raise ExcelWriteError(f"保存Excel文件失败: {save_error}")

    @handle_excel_operation
    def update_parameter_sheet(
        self,
        file_path: Path,
        sheet_name: str,
        parameter_data: Dict[str, List[str]],
        create_named_ranges: bool = True
    ) -> bool:
        """
        更新参数表工作表并创建命名区域
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（如"参数表"）
            parameter_data: 参数数据 {参数类型: [参数值列表]}
            create_named_ranges: 是否创建命名区域
            
        Returns:
            bool: 是否更新成功
        """
        try:
            logger.info(f"更新参数表: {file_path} -> {sheet_name}")
            
            # 加载工作簿
            wb = load_workbook(file_path)
            
            # 获取或创建工作表
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 清空现有数据（保留格式和命名区域）
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)
            
            # 创建居中对齐样式
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入参数数据
            for col_idx, (param_type, param_values) in enumerate(parameter_data.items(), 1):
                # 写入表头
                cell = ws.cell(row=1, column=col_idx, value=param_type)
                cell.alignment = center_alignment
                
                # 写入参数值
                for row_idx, value in enumerate(param_values, 2):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_alignment
                
                # 创建命名区域
                if create_named_ranges and param_values:
                    self._create_named_range(wb, sheet_name, param_type, col_idx)
            
            # 保存工作簿
            wb.save(file_path)
            logger.info(f"参数表更新成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"更新参数表失败: {file_path}", exc_info=True)
            raise ExcelWriteError(f"更新参数表失败: {file_path}", e)
    
    def _create_named_range(
        self,
        wb: Workbook,
        sheet_name: str,
        param_type: str,
        col_idx: int
    ):
        """
        创建命名区域
        
        Args:
            wb: 工作簿对象
            sheet_name: 工作表名称
            param_type: 参数类型
            col_idx: 列索引
        """
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.workbook.defined_name import DefinedName
            
            range_name = f"{param_type}List"
            col_letter = get_column_letter(col_idx)
            
            # 动态范围公式
            dynamic_range = f"OFFSET({sheet_name}!${col_letter}$2,0,0,COUNTA({sheet_name}!${col_letter}:${col_letter})-1,1)"
            
            # 删除已存在的同名区域
            if range_name in wb.defined_names:
                del wb.defined_names[range_name]
            
            # 创建新的命名区域
            wb.defined_names[range_name] = DefinedName(
                name=range_name,
                attr_text=dynamic_range
            )
            logger.debug(f"创建命名区域: {range_name} = {dynamic_range}")
            
        except Exception as e:
            logger.warning(f"创建命名区域失败 {param_type}: {e}")

