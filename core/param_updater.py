import pandas as pd
import os
import json
import core.data_reader as reader
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from core.config import PARAM_FILE, TARGET_PATH, MAPPINGS_FILE, VARIENT_DATA_FILE, VARIENT_MAPPINGS_FILE
# from core.config import CURRENT_PARAM_NAMES as PARAM_NAMES
from core.sentence_generator_manager import SentenceGeneratorManager
from openpyxl.utils import get_column_letter

class ParamUpdater:
    """参数表更新器 - 用于更新Excel中的数据验证和命名区域"""
    def __init__(self, engine_type=None):
        from core.config import ENGINE_TYPE
        self.engine_type = engine_type or ENGINE_TYPE
        self.manager = SentenceGeneratorManager(self.engine_type)

    def get_validation_categories(self):
        """获取需要验证的参数类别"""
        return self.manager.get_translatable_param_names()

    def generate_mappings_from_sheets(self, sheets, allowed_sheets=None):
        """
        通用的映射生成函数

        参数:
        sheets: 字典，工作表名 -> DataFrame
        allowed_sheets: 允许处理的工作表列表，None表示处理所有工作表

        返回:
        映射字典
        """
        mappings = {}

        # 确定要处理的工作表
        sheets_to_process = sheets.keys() if allowed_sheets is None else allowed_sheets

        for param_name in sheets_to_process:
            if param_name not in sheets:
                print(f"警告: {param_name} 工作表不存在")
                continue

            df = sheets[param_name]

            # 检查是否有足够的列
            if len(df.columns) < 2:
                print(f"警告: {param_name} 表没有足够的列")
                continue

            # 检查是否包含所需的列
            if "ExcelParam" not in df.columns or "ScenarioParam" not in df.columns:
                print(f"警告: {param_name} 表不包含 'ExcelParam' 或 'ScenarioParam' 列")
                continue

            # 构建映射
            param_mapping = {}
            for _, row in df.iterrows():
                excel_param = row["ExcelParam"]
                scenario_param = row["ScenarioParam"]

                if pd.notna(excel_param) and pd.notna(scenario_param):
                    param_mapping[str(excel_param)] = str(scenario_param)

            mappings[param_name] = param_mapping
            print(f"已处理 {param_name} 表，生成 {len(param_mapping)} 个映射")

        return mappings


    def save_mappings_to_file(self, mappings, file_path, variable_name="PARAM_MAPPINGS"):
        """
        通用的映射保存函数

        参数:
        mappings: 映射字典
        file_path: 保存路径
        variable_name: 变量名
        """
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("# 自动生成的参数映射\n")
                f.write("# 请不要手动编辑此文件\n\n")
                f.write(f"{variable_name} = ")
                f.write(repr(mappings))
            print(f"映射已保存到 Python 模块: {file_path}")
            return True
        except Exception as e:
            print(f"保存映射文件失败: {e}")
            return False


    def generate_base_param_mappings(self):
        """生成基础参数映射（param_data.xlsx）"""
        try:
            param_sheets = pd.read_excel(PARAM_FILE,sheet_name=None)
        except Exception as e:
            print(f"读取基础参数文件失败: {e}")
            return {}

        # 使用通用函数生成映射
        mappings = self.generate_mappings_from_sheets(param_sheets)

        # 保存映射
        self.save_mappings_to_file(mappings, MAPPINGS_FILE, "PARAM_MAPPINGS")

        return mappings

    def generate_varient_param_mappings(self):
        """生成差分参数映射（varient_data.xlsx）"""
        if not os.path.exists(VARIENT_DATA_FILE):
            print(f"差分参数文件不存在: {VARIENT_DATA_FILE}")
            return {}, []

        try:
            # 读取所有工作表（角色名作为工作表名）
            varient_sheets = pd.read_excel(VARIENT_DATA_FILE, sheet_name=None)
        except Exception as e:
            print(f"读取差分参数文件失败: {e}")
            return {}, []

        # 过滤掉不需要的工作表
        sheets_to_process = []
        for sheet_name in varient_sheets.keys():
            if sheet_name != "基础参数模板":  # 跳过基础参数模板
                sheets_to_process.append(sheet_name)

        # 使用通用函数生成映射（处理所有工作表）
        mappings = self.generate_mappings_from_sheets(varient_sheets)

        # 提取所有差分参数名（用于参数表更新）
        all_varient_params = set()
        for character_name, mapping in mappings.items():
            all_varient_params.update(mapping.keys())

        # 保存差分映射到单独文件
        self.save_mappings_to_file(mappings, VARIENT_MAPPINGS_FILE, "VARIENT_MAPPINGS")

        return mappings, sorted(list(all_varient_params))

    def collect_validation_data(self):
        """收集所有验证数据（基础参数 + 差分参数）"""
        validation_data = {}
        param_categories = self.get_validation_categories()

        print(f"需要验证的参数: {param_categories}")

        # 1. 收集基础参数
        try:
            base_sheets = pd.read_excel(PARAM_FILE, sheet_name=None)
        except Exception as e:
            print(f"读取基础参数文件失败: {e}")
            return {}
        
        missing_sheets = []
        collected_sheets = {}  # 记录每个工作表收集的参数
        sheet_param_map = {}   # 记录工作表到参数的映射

        # 首先收集所有参数到对应的工作表
        for param_name in param_categories:
            # 获取参数的所有翻译类型
            translate_types = self.manager.get_translate_types(param_name)

            # 收集所有相关工作表的参数
            for sheet_name in translate_types:
                if sheet_name in base_sheets:
                    df = base_sheets[sheet_name]
                    params = reader.read_column_from_dataframe(df, "ExcelParam")
                    if params:
                        # 初始化工作表记录
                        if sheet_name not in collected_sheets:
                            collected_sheets[sheet_name] = set()
                        if sheet_name not in sheet_param_map:
                            sheet_param_map[sheet_name] = []
                        
                        # 添加参数到工作表
                        collected_sheets[sheet_name].update(params)
                        sheet_param_map[sheet_name].append(param_name)
                        
                        # 添加到验证数据

                        validation_data[param_name] = params
                else:
                    if sheet_name not in missing_sheets:
                        missing_sheets.append(sheet_name)

        # 将集合转换为排序列表
        for param_name in validation_data:
            validation_data[param_name] = list(validation_data[param_name])

        # 输出工作表收集总结
        for sheet_name, params_set in collected_sheets.items():
            param_names = sheet_param_map.get(sheet_name, [])
            print(f"从工作表 '{sheet_name}' 收集参数: {len(params_set)} 个 (用于: {', '.join(param_names)})")

        # 输出成功收集的参数（按工作表分组，避免重复）
        collected_summary = []
        for sheet_name, params_set in collected_sheets.items():
            count = len(params_set)
            collected_summary.append(f"{sheet_name}({count})")

        if collected_summary:
            print(f"成功收集参数: {', '.join(collected_summary)}")
        
        if missing_sheets:
            print(f"警告: 以下工作表不存在: {', '.join(missing_sheets)}")

        # 2. 收集差分参数
        varient_mappings, varient_params = self.generate_varient_param_mappings()
        if varient_params:
            # 将差分参数合并到Varient列
            if "Varient" in validation_data:
                # 合并去重
                existing_varients = set(validation_data["Varient"])
                combined_varients = existing_varients.union(set(varient_params))
                validation_data["Varient"] = sorted(list(combined_varients))
            else:
                validation_data["Varient"] = varient_params
            
            # 如果有差分参数，添加到总结中
            if varient_params and "Varient" not in [item.split('(')[0] for item in collected_summary]:
                collected_summary.append(f"Varient({len(varient_params)})")
                print(f"成功收集参数: {', '.join(collected_summary)}")

        return validation_data

    def validation_update(self):
        """更新表格的数据验证参数并创建/更新命名区域"""
        # 使用新的数据收集函数获取所有验证数据（基础参数 + 差分参数）
        validation_data = self.collect_validation_data()

        if not validation_data:
            print("错误: 没有收集到验证数据")
            return False

        # 确保目标路径存在
        if not os.path.exists(TARGET_PATH):
            print(f"错误: 目标路径不存在: {TARGET_PATH}")
            return False

        # 处理每个Excel文件
        files = os.listdir(TARGET_PATH)
        excel_files = [f for f in files if f.endswith(".xlsx") and not f.startswith("~")]

        if not excel_files:
            print(f"警告: 在 {TARGET_PATH} 中没有找到Excel文件")
            return True

        success_count = 0
        updated_columns = []
        updated_ranges = []

        for file in excel_files:
            file_path = os.path.join(TARGET_PATH, file)

            try:
                # 加载工作簿
                wb = load_workbook(file_path)

                # 检查是否存在"参数表"工作表
                if "参数表" not in wb.sheetnames:
                    print(f"警告: {file} 中没有'参数表'工作表，跳过")
                    continue

                validation_ws = wb["参数表"]

                # 创建居中对齐样式
                center_alignment = Alignment(horizontal='center', vertical='center')

                # 获取需要验证的参数类别
                param_categories = self.get_validation_categories()
                # print(f"将按照以下顺序更新参数表: {param_categories}")

                created_ranges = set()  # 记录已经创建的命名区域
                # 按照固定顺序更新参数表（使用PARAM_NAMES的顺序）
                used_translate_types = set()
                unique_params = []

                for param_name in param_categories:
                    translate_types = self.manager.get_translate_types(param_name)
                    translate_type = translate_types[0] if translate_types else param_name
                    
                    if translate_type not in used_translate_types:
                        used_translate_types.add(translate_type)
                        unique_params.append((translate_type, param_name))

                for col_idx, param_tuple in enumerate(unique_params, 1):
                    translate_type, original_param_name = param_tuple
                    
                    # 检查这个参数是否有数据
                    params = validation_data.get(original_param_name, [])

                    # 检查是否需要更新这一列
                    needs_update = False

                    # 检查表头是否匹配
                    current_header = validation_ws.cell(row=1, column=col_idx).value
                    if current_header != translate_type:
                        needs_update = True
                        print(f"表头不匹配: 当前'{current_header}'，期望'{translate_type}'")
                    else:
                        # 检查参数内容是否匹配
                        current_params = []
                        row_idx = 2
                        while row_idx <= validation_ws.max_row:
                            cell_value = validation_ws.cell(row=row_idx, column=col_idx).value
                            if cell_value is None:
                                break
                            current_params.append(str(cell_value))
                            row_idx += 1

                        # 如果参数数量或内容有变化，需要更新
                        if current_params != [str(p) for p in params]:
                            needs_update = True
                            print(f"参数内容变化: 当前{len(current_params)}个，期望{len(params)}个")

                    # 更新列数据
                    if needs_update:
                        for row_idx in range(1, validation_ws.max_row + 1):
                            validation_ws.cell(row=row_idx, column=col_idx).value = None
                        
                        # 设置表头
                        validation_ws.cell(row=1, column=col_idx, value=translate_type).alignment = center_alignment
                        
                        for row_idx, param_value in enumerate(params, 2):
                            validation_ws.cell(row=row_idx, column=col_idx, value=param_value).alignment = center_alignment
                        
                        updated_columns.append(translate_type)

                    # 基于 translate_type 创建命名区域
                    translate_types = self.manager.get_translate_types(original_param_name)
                    if translate_types:
                        sheet_name = translate_types[0]
                        range_name = f"{sheet_name}List"
                        
                        if range_name in created_ranges:
                            continue
                        
                        created_ranges.add(range_name)
                        
                        col_letter = get_column_letter(col_idx)
                        expected_dynamic_range = f"OFFSET(参数表!${col_letter}$2,0,0,COUNTA(参数表!${col_letter}:${col_letter})-1,1)"

                        needs_range_update = False
                        if range_name in wb.defined_names:
                            existing_range = wb.defined_names[range_name].attr_text
                            if existing_range != expected_dynamic_range:
                                needs_range_update = True
                        else:
                            needs_range_update = True

                        if needs_range_update:
                            try:
                                from openpyxl.workbook.defined_name import DefinedName
                                if range_name in wb.defined_names:
                                    del wb.defined_names[range_name]
                                
                                wb.defined_names[range_name] = DefinedName(name=range_name, attr_text=expected_dynamic_range)
                                updated_ranges.append(range_name)
                            except Exception as e:
                                print(f"处理动态命名区域 {range_name} 时发生错误: {e}")

                # 保存工作簿
                try:
                    wb.save(file_path)
                    print(f"成功保存文件: {file}")
                    success_count += 1

                    # 输出更新总结
                    if updated_columns:
                        print(f"文件 {file}: 更新了 {len(updated_columns)} 列 - {', '.join(updated_columns)}")
                    if updated_ranges:
                        print(f"文件 {file}: 更新了 {len(updated_ranges)} 个命名区域")
                    if not updated_columns and not updated_ranges:
                        print(f"文件 {file}: 无需更新")

                except Exception as e:
                    print(f"保存失败: {e}")
                    # 尝试关闭工作簿以防止文件锁定
                    try:
                        wb.close()
                    except:
                        pass

            except Exception as e:
                print(f"处理文件 {file} 时发生错误: {e}")
                # 尝试关闭工作簿以防止文件锁定
                try:
                    wb.close()
                except:
                    pass

        print(f"处理完成，成功更新 {success_count}/{len(excel_files)} 个文件")
        return success_count > 0