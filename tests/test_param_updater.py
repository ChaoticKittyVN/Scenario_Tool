"""
测试 ParamUpdater 类
"""
import pytest
import pandas as pd
from pathlib import Path
from unittest.mock import Mock, patch
from update_param import ParamUpdater
from core.config_manager import AppConfig


class TestParamUpdater:
    """测试 ParamUpdater 类"""

    @pytest.fixture
    def mock_config(self, tmp_path):
        """创建模拟的配置对象"""
        config = Mock(spec=AppConfig)
        config.engine = Mock()
        config.engine.engine_type = "renpy"
        config.paths = Mock()
        config.paths.param_config_dir = tmp_path / "param_config"
        config.paths.input_dir = tmp_path / "input"

        # 创建必要的目录
        config.paths.param_config_dir.mkdir(parents=True, exist_ok=True)
        config.paths.input_dir.mkdir(parents=True, exist_ok=True)

        return config

    @pytest.fixture
    def mock_param_excel(self, tmp_path):
        """创建模拟的参数 Excel 文件"""
        param_file = tmp_path / "param_config" / "param_data_renpy.xlsx"
        param_file.parent.mkdir(parents=True, exist_ok=True)

        # 创建多个工作表
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            # Music 工作表
            df_music = pd.DataFrame({
                'ExcelParam': ['音乐1', '音乐2', '背景音乐'],
                'ScenarioParam': ['music1', 'music2', 'bgm_main']
            })
            df_music.to_excel(writer, sheet_name='Music', index=False)

            # Speaker 工作表
            df_speaker = pd.DataFrame({
                'ExcelParam': ['角色A', '角色B'],
                'ScenarioParam': ['character_a', 'character_b']
            })
            df_speaker.to_excel(writer, sheet_name='Speaker', index=False)

            # Background 工作表
            df_bg = pd.DataFrame({
                'ExcelParam': ['背景1', '背景2'],
                'ScenarioParam': ['bg_1', 'bg_2']
            })
            df_bg.to_excel(writer, sheet_name='Background', index=False)

            # Varient 工作表
            df_varient = pd.DataFrame({
                'ExcelParam': ['差分1', '差分2'],
                'ScenarioParam': ['variant_1', 'variant_2']
            })
            df_varient.to_excel(writer, sheet_name='Varient', index=False)

        return param_file

    @pytest.fixture
    def mock_varient_excel(self, tmp_path):
        """创建模拟的差分参数 Excel 文件"""
        varient_file = tmp_path / "param_config" / "varient_data.xlsx"

        with pd.ExcelWriter(varient_file, engine='openpyxl') as writer:
            # 角色A 工作表
            df_role_a = pd.DataFrame({
                'ExcelParam': ['开心', '难过'],
                'ScenarioParam': ['happy', 'sad']
            })
            df_role_a.to_excel(writer, sheet_name='角色A', index=False)

            # 角色B 工作表
            df_role_b = pd.DataFrame({
                'ExcelParam': ['生气', '惊讶'],
                'ScenarioParam': ['angry', 'surprised']
            })
            df_role_b.to_excel(writer, sheet_name='角色B', index=False)

            # 模板工作表（应该被处理但可能是空的）
            df_template = pd.DataFrame({
                'ExcelParam': [],
                'ScenarioParam': []
            })
            df_template.to_excel(writer, sheet_name='参数表模板', index=False)

        return varient_file

    @pytest.fixture
    def updater(self, mock_config):
        """创建 ParamUpdater 实例"""
        return ParamUpdater(mock_config)

    def test_init(self, updater, mock_config):
        """测试初始化"""
        assert updater.config == mock_config
        assert updater.engine_type == "renpy"

    def test_read_param_file_success(self, updater, mock_param_excel):
        """测试成功读取参数文件"""
        mappings = updater.read_param_file(mock_param_excel)

        # 验证读取到的映射
        assert 'Music' in mappings
        assert 'Speaker' in mappings
        assert 'Background' in mappings
        assert 'Varient' in mappings

        # 验证映射内容
        assert mappings['Music']['音乐1'] == 'music1'
        assert mappings['Speaker']['角色A'] == 'character_a'
        assert mappings['Background']['背景1'] == 'bg_1'

    def test_read_param_file_not_exist(self, updater, tmp_path):
        """测试读取不存在的文件"""
        non_existent = tmp_path / "nonexistent.xlsx"
        mappings = updater.read_param_file(non_existent)

        assert mappings == {}

    def test_read_param_file_skip_template(self, updater, tmp_path):
        """测试跳过模板工作表"""
        param_file = tmp_path / "test_param.xlsx"

        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            # 正常工作表
            df_normal = pd.DataFrame({
                'ExcelParam': ['参数1'],
                'ScenarioParam': ['param1']
            })
            df_normal.to_excel(writer, sheet_name='Normal', index=False)

            # 模板工作表
            df_template = pd.DataFrame({
                'ExcelParam': ['模板参数'],
                'ScenarioParam': ['template_param']
            })
            df_template.to_excel(writer, sheet_name='参数表模板', index=False)

        # skip_template=True 时应该跳过模板
        mappings = updater.read_param_file(param_file, skip_template=True)
        assert 'Normal' in mappings
        assert '参数表模板' not in mappings

        # skip_template=False 时应该包含模板
        mappings = updater.read_param_file(param_file, skip_template=False)
        assert 'Normal' in mappings
        assert '参数表模板' in mappings

    def test_read_param_file_missing_columns(self, updater, tmp_path):
        """测试缺少必需列的工作表"""
        param_file = tmp_path / "test_param.xlsx"

        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            # 缺少 ScenarioParam 列
            df_missing = pd.DataFrame({
                'ExcelParam': ['参数1'],
                'WrongColumn': ['wrong']
            })
            df_missing.to_excel(writer, sheet_name='MissingColumn', index=False)

        mappings = updater.read_param_file(param_file)

        # 应该跳过缺少必需列的工作表
        assert 'MissingColumn' not in mappings

    def test_collect_validation_data(self, updater, mock_param_excel):
        """测试收集验证数据"""
        validation_data = updater.collect_validation_data(mock_param_excel)

        # 验证收集到的数据
        assert 'Music' in validation_data
        assert 'Speaker' in validation_data
        assert 'Background' in validation_data
        assert 'Varient' in validation_data

        # 验证数据内容
        assert '音乐1' in validation_data['Music']
        assert '音乐2' in validation_data['Music']
        assert '角色A' in validation_data['Speaker']
        assert '背景1' in validation_data['Background']

    def test_collect_validation_data_with_varient(self, updater, mock_param_excel, mock_varient_excel):
        """测试收集验证数据（包含差分参数）"""
        validation_data = updater.collect_validation_data(mock_param_excel, mock_varient_excel)

        # 验证差分参数被合并到 Varient 列
        assert 'Varient' in validation_data

        # 应该包含基础差分参数
        assert '差分1' in validation_data['Varient']
        assert '差分2' in validation_data['Varient']

        # 应该包含角色特定的差分参数
        assert '开心' in validation_data['Varient']
        assert '难过' in validation_data['Varient']
        assert '生气' in validation_data['Varient']
        assert '惊讶' in validation_data['Varient']

        # 验证去重和排序
        varient_list = validation_data['Varient']
        assert len(varient_list) == len(set(varient_list))  # 无重复
        assert varient_list == sorted(varient_list)  # 已排序

    def test_collect_validation_data_file_not_exist(self, updater, tmp_path):
        """测试文件不存在时的行为"""
        non_existent = tmp_path / "nonexistent.xlsx"
        validation_data = updater.collect_validation_data(non_existent)

        assert validation_data == {}

    def test_collect_validation_data_empty_values(self, updater, tmp_path):
        """测试处理空值和空字符串"""
        param_file = tmp_path / "test_param.xlsx"

        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['参数1', '', None, '  ', '参数2'],
                'ScenarioParam': ['param1', 'empty', 'none', 'spaces', 'param2']
            })
            df.to_excel(writer, sheet_name='Test', index=False)

        validation_data = updater.collect_validation_data(param_file)

        # 应该只包含非空的参数
        assert 'Test' in validation_data
        assert '参数1' in validation_data['Test']
        assert '参数2' in validation_data['Test']
        assert '' not in validation_data['Test']
        assert '  ' not in validation_data['Test']

    def test_generate_mappings_file(self, updater, tmp_path):
        """测试生成映射文件"""
        mappings = {
            'Music': {'音乐1': 'music1', '音乐2': 'music2'},
            'Speaker': {'角色A': 'character_a'}
        }

        output_file = tmp_path / "param_mappings.py"
        updater.generate_mappings_file(mappings, output_file)

        # 验证文件被创建
        assert output_file.exists()

        # 验证文件内容
        content = output_file.read_text(encoding='utf-8')
        assert 'PARAM_MAPPINGS' in content
        assert 'Music' in content
        assert 'Speaker' in content
        assert 'music1' in content
        assert 'character_a' in content

    def test_generate_mappings_file_varient(self, updater, tmp_path):
        """测试生成差分映射文件"""
        varient_mappings = {
            '角色A': {'开心': 'happy'},
            '角色B': {'生气': 'angry'}
        }

        output_file = tmp_path / "varient_mappings.py"
        updater.generate_mappings_file(varient_mappings, output_file)

        # 验证文件被创建
        assert output_file.exists()

        # 验证文件内容
        content = output_file.read_text(encoding='utf-8')
        assert 'VARIENT_MAPPINGS' in content
        assert '角色A' in content
        assert 'happy' in content

    def test_update_scenario_param_sheets_no_validation_data(self, updater):
        """测试没有验证数据时的行为"""
        result = updater.update_scenario_param_sheets({})
        assert result is False

    def test_update_scenario_param_sheets_input_dir_not_exist(self, updater, tmp_path):
        """测试输入目录不存在时的行为"""
        updater.config.paths.input_dir = tmp_path / "nonexistent"
        validation_data = {'Music': ['音乐1']}

        result = updater.update_scenario_param_sheets(validation_data)
        assert result is False

    def test_update_scenario_param_sheets_no_excel_files(self, updater):
        """测试没有 Excel 文件时的行为"""
        validation_data = {'Music': ['音乐1']}

        result = updater.update_scenario_param_sheets(validation_data)
        # 没有文件时应该返回 True（不算错误）
        assert result is True

    def test_update_scenario_param_sheets_with_param_sheet(self, updater):
        """测试更新包含参数表的 Excel 文件"""
        from openpyxl import Workbook

        # 创建一个包含"参数表"的 Excel 文件
        excel_file = updater.config.paths.input_dir / "test_scenario.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "参数表"

        # 添加表头和旧数据
        ws['A1'] = 'Music'
        ws['B1'] = 'Speaker'
        ws['A2'] = '旧音乐1'
        ws['A3'] = '旧音乐2'
        ws['B2'] = '旧角色A'

        wb.save(excel_file)
        wb.close()

        # 准备新的验证数据
        validation_data = {
            'Music': ['音乐1', '音乐2', '音乐3'],
            'Speaker': ['角色A', '角色B']
        }

        # 执行更新
        result = updater.update_scenario_param_sheets(validation_data)

        # 验证结果
        assert result is True

        # 验证文件已更新
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb['参数表']

        # 验证 Music 列已更新
        music_values = []
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if value:
                music_values.append(value)

        assert len(music_values) == 3
        assert '音乐1' in music_values
        assert '音乐3' in music_values

        wb.close()

    def test_update_scenario_param_sheets_without_param_sheet(self, updater):
        """测试处理没有参数表的 Excel 文件"""
        # 创建一个没有"参数表"的 Excel 文件
        excel_file = updater.config.paths.input_dir / "no_param_sheet.xlsx"

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df = pd.DataFrame({'Data': [1, 2, 3]})
            df.to_excel(writer, sheet_name='Sheet1', index=False)

        validation_data = {'Music': ['音乐1']}

        # 应该跳过这个文件，但不报错
        result = updater.update_scenario_param_sheets(validation_data)
        assert result is False  # 没有成功更新任何文件

    def test_update_scenario_param_sheets_skip_temp_files(self, updater):
        """测试跳过临时文件（以 ~ 开头）"""
        from openpyxl import Workbook

        # 创建临时文件
        temp_file = updater.config.paths.input_dir / "~temp.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "参数表"
        ws['A1'] = 'Music'
        wb.save(temp_file)
        wb.close()

        validation_data = {'Music': ['音乐1']}

        # 应该跳过临时文件
        result = updater.update_scenario_param_sheets(validation_data)
        assert result is True  # 没有找到非临时文件，返回 True

    def test_update_scenario_param_sheets_no_changes_needed(self, updater):
        """测试当参数表已经是最新时的行为"""
        from openpyxl import Workbook

        # 创建一个已经包含正确数据的 Excel 文件
        excel_file = updater.config.paths.input_dir / "up_to_date.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "参数表"

        # 添加表头和数据（与验证数据相同）
        ws['A1'] = 'Music'
        ws['B1'] = 'Speaker'
        ws['A2'] = '音乐1'
        ws['A3'] = '音乐2'
        ws['B2'] = '角色A'

        wb.save(excel_file)
        wb.close()

        # 使用相同的数据
        validation_data = {
            'Music': ['音乐1', '音乐2'],
            'Speaker': ['角色A']
        }

        # 执行更新
        result = updater.update_scenario_param_sheets(validation_data)

        # 应该检测到需要创建命名区域，所以返回 True
        assert result is True

    def test_update_scenario_param_sheets_with_named_ranges(self, updater):
        """测试创建命名区域"""
        from openpyxl import Workbook, load_workbook

        # 创建 Excel 文件
        excel_file = updater.config.paths.input_dir / "with_ranges.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "参数表"
        ws['A1'] = 'Music'
        ws['B1'] = 'Speaker'
        ws['A2'] = '音乐1'
        ws['B2'] = '角色A'

        wb.save(excel_file)
        wb.close()

        validation_data = {
            'Music': ['音乐1', '音乐2'],
            'Speaker': ['角色A', '角色B']
        }

        # 执行更新
        result = updater.update_scenario_param_sheets(validation_data)
        assert result is True

        # 验证命名区域已创建
        wb = load_workbook(excel_file)

        # 检查 MusicList 命名区域
        assert 'MusicList' in wb.defined_names
        music_range = wb.defined_names['MusicList']
        assert 'OFFSET' in music_range.attr_text
        assert 'COUNTA' in music_range.attr_text

        # 检查 SpeakerList 命名区域
        assert 'SpeakerList' in wb.defined_names

        wb.close()

    def test_update_scenario_param_sheets_multiple_files(self, updater):
        """测试同时处理多个 Excel 文件"""
        from openpyxl import Workbook, load_workbook

        # 创建多个 Excel 文件
        for i in range(3):
            excel_file = updater.config.paths.input_dir / f"scenario_{i}.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "参数表"
            ws['A1'] = 'Music'
            ws['A2'] = '旧音乐'

            wb.save(excel_file)
            wb.close()

        validation_data = {
            'Music': ['音乐1', '音乐2']
        }

        # 执行更新
        result = updater.update_scenario_param_sheets(validation_data)
        assert result is True

        # 验证所有文件都已更新
        for i in range(3):
            excel_file = updater.config.paths.input_dir / f"scenario_{i}.xlsx"
            wb = load_workbook(excel_file)
            ws = wb['参数表']

            # 验证数据已更新
            music_col_values = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row=row, column=1).value
                if value:
                    music_col_values.append(value)

            assert len(music_col_values) == 2
            assert '音乐1' in music_col_values
            wb.close()


class TestParamUpdaterIntegration:
    """集成测试：测试完整的参数更新流程"""

    @pytest.fixture
    def full_setup(self, tmp_path):
        """创建完整的测试环境"""
        # 创建配置
        config = Mock(spec=AppConfig)
        config.engine = Mock()
        config.engine.engine_type = "renpy"
        config.paths = Mock()
        config.paths.param_config_dir = tmp_path / "param_config"
        config.paths.input_dir = tmp_path / "input"

        # 创建目录
        config.paths.param_config_dir.mkdir(parents=True)
        config.paths.input_dir.mkdir(parents=True)

        # 创建参数文件
        param_file = config.paths.param_config_dir / "param_data_renpy.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df_music = pd.DataFrame({
                'ExcelParam': ['音乐1', '音乐2'],
                'ScenarioParam': ['music1', 'music2']
            })
            df_music.to_excel(writer, sheet_name='Music', index=False)

        return config, param_file

    def test_full_workflow(self, full_setup):
        """测试完整的工作流程"""
        config, param_file = full_setup
        updater = ParamUpdater(config)

        # 1. 读取参数文件
        mappings = updater.read_param_file(param_file)
        assert 'Music' in mappings

        # 2. 生成映射文件
        output_file = config.paths.param_config_dir / "param_mappings.py"
        updater.generate_mappings_file(mappings, output_file)
        assert output_file.exists()

        # 3. 收集验证数据
        validation_data = updater.collect_validation_data(param_file)
        assert 'Music' in validation_data
        assert '音乐1' in validation_data['Music']


class TestExceptionHandling:
    """测试异常处理"""

    @pytest.fixture
    def mock_config(self, tmp_path):
        """创建模拟配置"""
        config = Mock(spec=AppConfig)
        config.engine = Mock()
        config.engine.engine_type = "renpy"
        config.paths = Mock()
        config.paths.param_config_dir = tmp_path / "param_config"
        config.paths.input_dir = tmp_path / "input"
        config.paths.param_config_dir.mkdir(parents=True, exist_ok=True)
        config.paths.input_dir.mkdir(parents=True, exist_ok=True)
        return config

    @pytest.fixture
    def updater(self, mock_config):
        """创建更新器实例"""
        return ParamUpdater(mock_config)

    def test_read_param_file_with_exception(self, updater, tmp_path):
        """测试读取参数文件时发生异常"""
        # 创建一个损坏的 Excel 文件（实际上是文本文件）
        bad_file = tmp_path / "bad.xlsx"
        bad_file.write_text("This is not an Excel file")

        result = updater.read_param_file(bad_file)

        # 应该返回空字典
        assert result == {}

    def test_generate_mappings_file_with_io_error(self, updater, tmp_path):
        """测试生成映射文件时发生 IO 错误"""
        mappings = {'Music': {'音乐1': 'music1'}}

        # 使用一个不存在的目录（不创建）
        bad_dir = tmp_path / "nonexistent" / "subdir"
        output_file = bad_dir / "mappings.py"

        # 应该捕获异常，不抛出
        updater.generate_mappings_file(mappings, output_file)

        # 文件不应该被创建
        assert not output_file.exists()

    def test_collect_validation_data_with_read_error(self, updater, tmp_path):
        """测试收集验证数据时读取失败"""
        # 创建一个损坏的文件
        bad_file = tmp_path / "bad.xlsx"
        bad_file.write_text("Not an Excel file")

        result = updater.collect_validation_data(bad_file)

        # 应该返回空字典
        assert result == {}

    def test_collect_validation_data_with_varient_read_error(self, updater, tmp_path):
        """测试收集验证数据时差分文件读取失败"""
        # 创建正常的参数文件
        param_file = tmp_path / "param.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['音乐1'],
                'ScenarioParam': ['music1']
            })
            df.to_excel(writer, sheet_name='Music', index=False)

        # 创建损坏的差分文件
        varient_file = tmp_path / "varient.xlsx"
        varient_file.write_text("Not an Excel file")

        # 应该只返回基础参数，忽略差分文件错误
        result = updater.collect_validation_data(param_file, varient_file)

        assert 'Music' in result
        assert '音乐1' in result['Music']


class TestUpdateMappingsMethod:
    """测试 update_mappings 方法"""

    @pytest.fixture
    def mock_config(self, tmp_path):
        """创建模拟配置"""
        config = Mock(spec=AppConfig)
        config.engine = Mock()
        config.engine.engine_type = "renpy"
        config.paths = Mock()
        config.paths.param_config_dir = tmp_path / "param_config"
        config.paths.input_dir = tmp_path / "input"
        config.paths.param_config_dir.mkdir(parents=True, exist_ok=True)
        config.paths.input_dir.mkdir(parents=True, exist_ok=True)
        return config

    @pytest.fixture
    def updater(self, mock_config):
        """创建更新器实例"""
        return ParamUpdater(mock_config)

    def test_update_mappings_param_file_not_exist(self, updater):
        """测试参数文件不存在时的行为"""
        result = updater.update_mappings()

        # 应该返回 False
        assert result is False

    def test_update_mappings_success_without_varient(self, updater, tmp_path):
        """测试成功更新映射（没有差分文件）"""
        # 创建参数文件
        param_file = updater.config.paths.param_config_dir / "param_data_renpy.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['音乐1', '音乐2'],
                'ScenarioParam': ['music1', 'music2']
            })
            df.to_excel(writer, sheet_name='Music', index=False)

        result = updater.update_mappings()

        # 应该成功
        assert result is True

        # 验证映射文件已创建
        mapping_file = updater.config.paths.param_config_dir / "param_mappings.py"
        assert mapping_file.exists()

    def test_update_mappings_success_with_varient(self, updater, tmp_path):
        """测试成功更新映射（有差分文件）"""
        # 创建参数文件
        param_file = updater.config.paths.param_config_dir / "param_data_renpy.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['音乐1'],
                'ScenarioParam': ['music1']
            })
            df.to_excel(writer, sheet_name='Music', index=False)

        # 创建差分文件
        varient_file = updater.config.paths.param_config_dir / "varient_data.xlsx"
        with pd.ExcelWriter(varient_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['开心', '难过'],
                'ScenarioParam': ['happy', 'sad']
            })
            df.to_excel(writer, sheet_name='角色A', index=False)

        result = updater.update_mappings()

        # 应该成功
        assert result is True

        # 验证两个映射文件都已创建
        mapping_file = updater.config.paths.param_config_dir / "param_mappings.py"
        varient_mapping_file = updater.config.paths.param_config_dir / "varient_mappings.py"
        assert mapping_file.exists()
        assert varient_mapping_file.exists()

    def test_update_mappings_empty_mappings(self, updater, tmp_path):
        """测试参数文件为空时的行为"""
        # 创建一个空的参数文件
        param_file = updater.config.paths.param_config_dir / "param_data_renpy.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df = pd.DataFrame({'A': []})
            df.to_excel(writer, sheet_name='Sheet1', index=False)

        result = updater.update_mappings()

        # 应该返回 False（没有读取到映射）
        assert result is False

    def test_update_mappings_with_scenario_param_sheets(self, updater, tmp_path):
        """测试更新映射并更新演出表格"""
        from openpyxl import Workbook

        # 创建参数文件
        param_file = updater.config.paths.param_config_dir / "param_data_renpy.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['音乐1', '音乐2'],
                'ScenarioParam': ['music1', 'music2']
            })
            df.to_excel(writer, sheet_name='Music', index=False)

        # 创建演出表格文件
        scenario_file = updater.config.paths.input_dir / "scenario.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "参数表"
        ws['A1'] = 'Music'
        wb.save(scenario_file)
        wb.close()

        result = updater.update_mappings()

        # 应该成功
        assert result is True


class TestEdgeCases:
    """测试边界情况"""

    @pytest.fixture
    def mock_config(self, tmp_path):
        """创建模拟配置"""
        config = Mock(spec=AppConfig)
        config.engine = Mock()
        config.engine.engine_type = "renpy"
        config.paths = Mock()
        config.paths.param_config_dir = tmp_path / "param_config"
        config.paths.input_dir = tmp_path / "input"
        config.paths.param_config_dir.mkdir(parents=True, exist_ok=True)
        config.paths.input_dir.mkdir(parents=True, exist_ok=True)
        return config

    @pytest.fixture
    def updater(self, mock_config):
        """创建更新器实例"""
        return ParamUpdater(mock_config)

    def test_update_scenario_param_sheets_named_range_already_correct(self, updater):
        """测试命名区域已经正确时不更新"""
        from openpyxl import Workbook
        from openpyxl.workbook.defined_name import DefinedName

        # 创建 Excel 文件
        excel_file = updater.config.paths.input_dir / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "参数表"
        ws['A1'] = 'Music'
        ws['A2'] = '音乐1'

        # 创建正确的命名区域
        correct_range = "OFFSET(参数表!$A$2,0,0,COUNTA(参数表!$A:$A)-1,1)"
        wb.defined_names['MusicList'] = DefinedName(
            name='MusicList',
            attr_text=correct_range
        )

        wb.save(excel_file)
        wb.close()

        validation_data = {'Music': ['音乐1']}

        # 执行更新
        result = updater.update_scenario_param_sheets(validation_data)

        # 应该返回 False（没有更新）
        assert result is False

    def test_collect_validation_data_merge_varient(self, updater, tmp_path):
        """测试合并差分参数到 Varient 列"""
        # 创建基础参数文件（包含 Varient）
        param_file = tmp_path / "param.xlsx"
        with pd.ExcelWriter(param_file, engine='openpyxl') as writer:
            df_music = pd.DataFrame({
                'ExcelParam': ['音乐1'],
                'ScenarioParam': ['music1']
            })
            df_music.to_excel(writer, sheet_name='Music', index=False)

            df_varient = pd.DataFrame({
                'ExcelParam': ['差分A', '差分B'],
                'ScenarioParam': ['var_a', 'var_b']
            })
            df_varient.to_excel(writer, sheet_name='Varient', index=False)

        # 创建差分文件
        varient_file = tmp_path / "varient.xlsx"
        with pd.ExcelWriter(varient_file, engine='openpyxl') as writer:
            df = pd.DataFrame({
                'ExcelParam': ['差分C', '差分A'],  # 差分A 重复
                'ScenarioParam': ['var_c', 'var_a']
            })
            df.to_excel(writer, sheet_name='角色A', index=False)

        result = updater.collect_validation_data(param_file, varient_file)

        # 验证 Varient 列合并且去重
        assert 'Varient' in result
        varient_list = result['Varient']
        assert '差分A' in varient_list
        assert '差分B' in varient_list
        assert '差分C' in varient_list
        # 应该去重，所以只有 3 个
        assert len(varient_list) == 3

    def test_generate_mappings_file_varient_variable_name(self, updater, tmp_path):
        """测试生成差分映射文件时使用正确的变量名"""
        mappings = {'角色A': {'开心': 'happy'}}
        output_file = tmp_path / "varient_mappings.py"

        updater.generate_mappings_file(mappings, output_file)

        # 读取文件内容
        content = output_file.read_text(encoding='utf-8')

        # 应该使用 VARIENT_MAPPINGS 变量名
        assert 'VARIENT_MAPPINGS' in content
        assert 'PARAM_MAPPINGS' not in content

    def test_generate_mappings_file_param_variable_name(self, updater, tmp_path):
        """测试生成参数映射文件时使用正确的变量名"""
        mappings = {'Music': {'音乐1': 'music1'}}
        output_file = tmp_path / "param_mappings.py"

        updater.generate_mappings_file(mappings, output_file)

        # 读取文件内容
        content = output_file.read_text(encoding='utf-8')

        # 应该使用 PARAM_MAPPINGS 变量名
        assert 'PARAM_MAPPINGS' in content
        assert 'VARIENT_MAPPINGS' not in content
